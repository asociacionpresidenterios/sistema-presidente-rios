import os
from datetime import date, datetime
from io import BytesIO

import qrcode

from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    jsonify,
    Response
)

from flask_sqlalchemy import SQLAlchemy
from openpyxl import load_workbook


# ============================================================
# CONFIGURACIÓN
# ============================================================

app = Flask(__name__)

app.config["SECRET_KEY"] = os.environ.get(
    "SECRET_KEY",
    "cambia-esta-clave-en-produccion"
)


# ============================================================
# BASE DE DATOS
# ============================================================

db_url = os.environ.get("DATABASE_URL")

if db_url:

    db_url = db_url.replace(
        "postgres://",
        "postgresql+psycopg2://",
        1
    )

    db_url = db_url.replace(
        "postgresql://",
        "postgresql+psycopg2://",
        1
    )


app.config["SQLALCHEMY_DATABASE_URI"] = (
    db_url or "sqlite:///jugadores.db"
)

app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)


# ============================================================
# ESTADOS PERMITIDOS
# ============================================================

ESTADOS_PERMITIDOS = {
    "Vigente",
    "Pendiente",
    "Suspendido",
    "Inhabilitado"
}


def normalizar_estado(estado):

    if not estado:
        return "Vigente"

    estado = str(estado).strip()

    if estado not in ESTADOS_PERMITIDOS:
        return "Vigente"

    return estado


# ============================================================
# MODELO CLUB
# ============================================================

class Club(db.Model):

    id = db.Column(
        db.Integer,
        primary_key=True
    )

    nombre = db.Column(
        db.String(120),
        unique=True,
        nullable=False
    )

    activo = db.Column(
        db.Boolean,
        nullable=False,
        default=True
    )


# ============================================================
# MODELO SERIE
# ============================================================

class Serie(db.Model):

    id = db.Column(
        db.Integer,
        primary_key=True
    )

    nombre = db.Column(
        db.String(80),
        unique=True,
        nullable=False
    )

    activo = db.Column(
        db.Boolean,
        nullable=False,
        default=True
    )


# ============================================================
# MODELO JUGADOR
# ============================================================

class Jugador(db.Model):

    id = db.Column(
        db.Integer,
        primary_key=True
    )

    rut = db.Column(
        db.String(20),
        unique=True,
        nullable=False,
        index=True
    )

    nombre_completo = db.Column(
        db.String(160),
        nullable=False
    )

    fecha_nacimiento = db.Column(
        db.Date,
        nullable=False
    )

    serie = db.Column(
        db.String(80),
        nullable=False
    )

    club = db.Column(
        db.String(120),
        nullable=False
    )

    foto = db.Column(
        db.LargeBinary,
        nullable=True
    )

    estado = db.Column(
        db.String(30),
        nullable=False,
        default="Vigente"
    )


# ============================================================
# MODELO REGISTRO DISCIPLINARIO
# ============================================================

class RegistroDisciplinario(db.Model):

    id = db.Column(
        db.Integer,
        primary_key=True
    )

    jugador_id = db.Column(
        db.Integer,
        db.ForeignKey("jugador.id"),
        nullable=False,
        index=True
    )

    fecha = db.Column(
        db.Date,
        nullable=False,
        default=date.today
    )

    tipo = db.Column(
        db.String(30),
        nullable=False
    )

    cantidad = db.Column(
        db.Integer,
        nullable=False,
        default=1
    )

    motivo = db.Column(
        db.String(255),
        nullable=True
    )

    campeonato = db.Column(
        db.String(120),
        nullable=True
    )

    observaciones = db.Column(
        db.Text,
        nullable=True
    )

    jugador = db.relationship(
        "Jugador",
        backref=db.backref(
            "registros_disciplinarios",
            lazy=True,
            cascade="all, delete-orphan"
        )
    )


# ============================================================
# MODELO GOLES
# ============================================================

class Gol(db.Model):

    id = db.Column(
        db.Integer,
        primary_key=True
    )

    jugador_id = db.Column(
        db.Integer,
        db.ForeignKey("jugador.id"),
        nullable=False,
        index=True
    )

    fecha = db.Column(
        db.Date,
        nullable=False,
        default=date.today
    )

    cantidad = db.Column(
        db.Integer,
        nullable=False,
        default=1
    )

    campeonato = db.Column(
        db.String(120),
        nullable=True
    )

    observaciones = db.Column(
        db.Text,
        nullable=True
    )

    jugador = db.relationship(
        "Jugador",
        backref=db.backref(
            "goles_registrados",
            lazy=True,
            cascade="all, delete-orphan"
        )
    )


# ============================================================
# MODELO CAMPEONATO
# ============================================================

class Campeonato(db.Model):

    id = db.Column(
        db.Integer,
        primary_key=True
    )

    nombre = db.Column(
        db.String(160),
        nullable=False
    )

    temporada = db.Column(
        db.String(20),
        nullable=False
    )

    serie = db.Column(
        db.String(80),
        nullable=False
    )

    fecha_inicio = db.Column(
        db.Date,
        nullable=True
    )

    fecha_termino = db.Column(
        db.Date,
        nullable=True
    )

    estado = db.Column(
        db.String(30),
        nullable=False,
        default="Activo"
    )

    descripcion = db.Column(
        db.Text,
        nullable=True
    )



# ============================================================
# CREACIÓN / ACTUALIZACIÓN SEGURA DE BASE DE DATOS
# ============================================================

def _columna_existe(tabla, columna):

    inspector = db.inspect(db.engine)

    return columna in {
        item["name"]
        for item in inspector.get_columns(tabla)
    }


def _agregar_columna_si_falta(tabla, columna, definicion):

    try:

        if _columna_existe(tabla, columna):
            return False

        dialecto = db.engine.dialect.name

        if dialecto == "postgresql":

            sql = (
                f'ALTER TABLE "{tabla}" '
                f'ADD COLUMN IF NOT EXISTS "{columna}" {definicion}'
            )

        elif dialecto == "sqlite":

            sql = (
                f'ALTER TABLE "{tabla}" '
                f'ADD COLUMN "{columna}" {definicion}'
            )

        else:

            sql = (
                f'ALTER TABLE "{tabla}" '
                f'ADD COLUMN "{columna}" {definicion}'
            )

        db.session.execute(db.text(sql))
        db.session.commit()

        print(
            f"[BD] Columna agregada: {tabla}.{columna}"
        )

        return True

    except Exception as error:

        db.session.rollback()

        print(
            f"[BD] No se pudo agregar {tabla}.{columna}: {repr(error)}"
        )

        return False


def preparar_base_datos():
    """
    Crea las tablas nuevas y realiza migraciones pequeñas y seguras
    para instalaciones existentes de SQLite/PostgreSQL.

    IMPORTANTE:
    db.create_all() NO modifica tablas que ya existen. Por eso aquí
    comprobamos columna por columna para evitar que una base antigua
    rompa las rutas de estadísticas disciplinarias.
    """

    try:
        db.create_all()
    except Exception as error:
        db.session.rollback()
        print("[BD] Error ejecutando db.create_all():", repr(error))
        return

    # --------------------------------------------------------
    # JUGADOR
    # --------------------------------------------------------

    _agregar_columna_si_falta(
        "jugador",
        "foto",
        "BYTEA" if db.engine.dialect.name == "postgresql" else "BLOB"
    )

    _agregar_columna_si_falta(
        "jugador",
        "estado",
        "VARCHAR(30) DEFAULT 'Vigente'"
    )

    try:

        db.session.execute(
            db.text(
                "UPDATE jugador "
                "SET estado = 'Vigente' "
                "WHERE estado IS NULL OR estado = ''"
            )
        )

        db.session.commit()

    except Exception as error:

        db.session.rollback()
        print("[BD] No se pudo normalizar estado de jugadores:", repr(error))

    # --------------------------------------------------------
    # REGISTRO DISCIPLINARIO
    # --------------------------------------------------------
    # Se revisan TODAS las columnas utilizadas por el modelo.
    # Esto es importante porque create_all() no actualiza una tabla
    # que ya existía en producción.

    columnas_disciplina = [
        ("jugador_id", "INTEGER"),
        ("fecha", "DATE"),
        ("tipo", "VARCHAR(30)"),
        ("cantidad", "INTEGER DEFAULT 1"),
        ("motivo", "VARCHAR(255)"),
        ("campeonato", "VARCHAR(120)"),
        ("observaciones", "TEXT")
    ]

    for nombre, definicion in columnas_disciplina:
        _agregar_columna_si_falta(
            "registro_disciplinario",
            nombre,
            definicion
        )

    # --------------------------------------------------------
    # GOLES
    # --------------------------------------------------------

    columnas_gol = [
        ("jugador_id", "INTEGER"),
        ("fecha", "DATE"),
        ("cantidad", "INTEGER DEFAULT 1"),
        ("campeonato", "VARCHAR(120)"),
        ("observaciones", "TEXT")
    ]

    for nombre, definicion in columnas_gol:
        _agregar_columna_si_falta(
            "gol",
            nombre,
            definicion
        )

    # --------------------------------------------------------
    # VERIFICACIÓN FINAL
    # --------------------------------------------------------

    try:

        inspector = db.inspect(db.engine)

        tablas = set(inspector.get_table_names())

        print("[BD] Tablas disponibles:", sorted(tablas))

        if "registro_disciplinario" in tablas:

            print(
                "[BD] Columnas registro_disciplinario:",
                [
                    c["name"]
                    for c in inspector.get_columns(
                        "registro_disciplinario"
                    )
                ]
            )

        if "gol" in tablas:

            print(
                "[BD] Columnas gol:",
                [
                    c["name"]
                    for c in inspector.get_columns("gol")
                ]
            )

    except Exception as error:

        print(
            "[BD] Advertencia verificando esquema final:",
            repr(error)
        )


with app.app_context():

    preparar_base_datos()


# ============================================================
# FUNCIONES AUXILIARES
# ============================================================

def normalizar_rut(rut):

    if rut is None:
        return ""

    rut = str(rut).strip().upper()

    rut = rut.replace(
        " ",
        ""
    )

    rut_limpio = (
        rut
        .replace(".", "")
        .replace("-", "")
    )

    if len(rut_limpio) < 2:
        return ""

    cuerpo = rut_limpio[:-1]

    dv = rut_limpio[-1]

    if not cuerpo.isdigit():
        return ""

    cuerpo_formateado = ""

    while len(cuerpo) > 3:

        cuerpo_formateado = (
            "."
            + cuerpo[-3:]
            + cuerpo_formateado
        )

        cuerpo = cuerpo[:-3]

    cuerpo_formateado = (
        cuerpo
        + cuerpo_formateado
    )

    return f"{cuerpo_formateado}-{dv}"


def validar_rut(rut):

    if not rut:
        return False

    rut_limpio = (
        rut
        .replace(".", "")
        .replace("-", "")
        .replace(" ", "")
        .upper()
    )

    if len(rut_limpio) < 2:
        return False

    cuerpo = rut_limpio[:-1]

    dv = rut_limpio[-1]

    if not cuerpo.isdigit():
        return False

    suma = 0

    multiplicador = 2

    for digito in reversed(cuerpo):

        suma += (
            int(digito)
            * multiplicador
        )

        multiplicador += 1

        if multiplicador > 7:
            multiplicador = 2

    resto = suma % 11

    resultado = 11 - resto

    if resultado == 11:

        dv_calculado = "0"

    elif resultado == 10:

        dv_calculado = "K"

    else:

        dv_calculado = str(resultado)

    return dv == dv_calculado


def convertir_fecha(valor):

    if valor is None:
        return None

    if isinstance(valor, datetime):

        return valor.date()

    if isinstance(valor, date):

        return valor

    if isinstance(valor, str):

        valor = valor.strip()

        formatos = [
            "%d/%m/%Y",
            "%d-%m-%Y",
            "%Y-%m-%d",
            "%d.%m.%Y"
        ]

        for formato in formatos:

            try:

                return datetime.strptime(
                    valor,
                    formato
                ).date()

            except ValueError:

                continue

    return None


def obtener_fotografia():

    archivo = request.files.get(
        "foto"
    )

    if not archivo:
        return None, None

    if not archivo.filename:
        return None, None

    contenido = archivo.read()

    if not contenido:

        return (
            None,
            "La fotografía seleccionada está vacía."
        )

    maximo = 5 * 1024 * 1024

    if len(contenido) > maximo:

        return (
            None,
            "La fotografía no puede superar los 5 MB."
        )

    tipo = (
        archivo.mimetype or ""
    ).lower()

    tipos_permitidos = {
        "image/jpeg",
        "image/png",
        "image/webp"
    }

    if tipo not in tipos_permitidos:

        return (
            None,
            "La fotografía debe ser JPG, PNG o WEBP."
        )

    return contenido, None


# ============================================================
# ESTADÍSTICAS DEPORTIVAS
# ============================================================

def obtener_amarillas(jugador_id):

    try:
        registros = (
            RegistroDisciplinario.query
            .filter_by(
                jugador_id=jugador_id,
                tipo="Amarilla"
            )
            .all()
        )

        return sum(
            registro.cantidad or 0
            for registro in registros
        )

    except Exception as error:

        db.session.rollback()
        print("Advertencia obteniendo amarillas:", error)
        return 0


def obtener_rojas(jugador_id):

    try:
        registros = (
            RegistroDisciplinario.query
            .filter_by(
                jugador_id=jugador_id,
                tipo="Roja"
            )
            .all()
        )

        return sum(
            registro.cantidad or 0
            for registro in registros
        )

    except Exception as error:

        db.session.rollback()
        print("Advertencia obteniendo rojas:", error)
        return 0


def obtener_goles(jugador_id):

    try:
        registros = (
            Gol.query
            .filter_by(
                jugador_id=jugador_id
            )
            .all()
        )

        return sum(
            registro.cantidad or 0
            for registro in registros
        )

    except Exception as error:

        db.session.rollback()
        print("Advertencia obteniendo goles:", error)
        return 0


def obtener_suspensiones(jugador_id):

    try:
        registros = (
            RegistroDisciplinario.query
            .filter_by(
                jugador_id=jugador_id,
                tipo="Suspension"
            )
            .all()
        )

        return sum(
            registro.cantidad or 0
            for registro in registros
        )

    except Exception as error:

        db.session.rollback()
        print("Advertencia obteniendo suspensiones:", error)
        return 0


def crear_suspension_por_acumulacion(
    jugador,
    campeonato=None
):

    amarillas = obtener_amarillas(
        jugador.id
    )

    suspensiones_correspondientes = (
        amarillas // 4
    )

    suspensiones_existentes = (
        obtener_suspensiones(
            jugador.id
        )
    )

    nuevas_suspensiones = (
        suspensiones_correspondientes
        - suspensiones_existentes
    )

    if nuevas_suspensiones <= 0:
        return 0

    jugador.estado = "Suspendido"

    for _ in range(
        nuevas_suspensiones
    ):

        suspension = RegistroDisciplinario(

            jugador_id=jugador.id,

            fecha=date.today(),

            tipo="Suspension",

            cantidad=1,

            motivo=(
                "Suspensión automática "
                "por acumulación de 4 "
                "tarjetas amarillas."
            ),

            campeonato=campeonato
        )

        db.session.add(
            suspension
        )

    return nuevas_suspensiones


# ============================================================
# DATOS PARA FORMULARIO DE JUGADORES
# ============================================================

def obtener_datos_formulario_jugador():

    clubes = Club.query.filter_by(
        activo=True
    ).order_by(
        Club.nombre
    ).all()

    series = Serie.query.filter_by(
        activo=True
    ).order_by(
        Serie.nombre
    ).all()

    return clubes, series


# ============================================================
# INICIO / LISTADO
# ============================================================

@app.route("/")
def index():

    q = request.args.get(
        "q",
        ""
    ).strip()

    query = Jugador.query

    if q:

        query = query.filter(
            db.or_(
                Jugador.rut.ilike(
                    f"%{q}%"
                ),

                Jugador.nombre_completo.ilike(
                    f"%{q}%"
                ),

                Jugador.club.ilike(
                    f"%{q}%"
                )
            )
        )

    jugadores = query.order_by(
        Jugador.nombre_completo
    ).all()

    return render_template(
        "index.html",
        jugadores=jugadores,
        q=q
    )


# ============================================================
# FICHA INDIVIDUAL
# ============================================================

@app.route(
    "/jugadores/<int:jugador_id>"
)
def ficha_jugador(jugador_id):

    jugador = db.get_or_404(
        Jugador,
        jugador_id
    )

    goles = obtener_goles(
        jugador.id
    )

    amarillas = obtener_amarillas(
        jugador.id
    )

    rojas = obtener_rojas(
        jugador.id
    )

    suspensiones = obtener_suspensiones(
        jugador.id
    )

    try:
        historial = (
            RegistroDisciplinario.query
            .filter_by(
                jugador_id=jugador.id
            )
            .order_by(
                RegistroDisciplinario.fecha.desc(),
                RegistroDisciplinario.id.desc()
            )
            .all()
        )
    except Exception as error:
        db.session.rollback()
        print("Advertencia obteniendo historial disciplinario:", error)
        historial = []

    try:
        historial_goles = (
            Gol.query
            .filter_by(
                jugador_id=jugador.id
            )
            .order_by(
                Gol.fecha.desc(),
                Gol.id.desc()
            )
            .all()
        )
    except Exception as error:
        db.session.rollback()
        print("Advertencia obteniendo historial de goles:", error)
        historial_goles = []

    return render_template(
        "jugador_detalle.html",
        jugador=jugador,
        goles=goles,
        amarillas=amarillas,
        rojas=rojas,
        suspensiones=suspensiones,
        historial=historial,
        historial_goles=historial_goles
    )


# ============================================================
# FOTOGRAFÍA
# ============================================================

@app.route(
    "/jugadores/<int:jugador_id>/foto"
)
def foto_jugador(jugador_id):

    jugador = db.get_or_404(
        Jugador,
        jugador_id
    )

    if not jugador.foto:

        return (
            "",
            404
        )

    return Response(
        jugador.foto,
        mimetype="image/jpeg"
    )


# ============================================================
# NUEVO JUGADOR
# ============================================================

@app.route(
    "/jugadores/nuevo",
    methods=["GET", "POST"]
)
def nuevo_jugador():

    clubes, series = obtener_datos_formulario_jugador()

    if request.method == "POST":

        rut = normalizar_rut(
            request.form.get(
                "rut",
                ""
            )
        )

        nombre = request.form.get(
            "nombre_completo",
            ""
        ).strip()

        fecha = request.form.get(
            "fecha_nacimiento",
            ""
        ).strip()

        serie = request.form.get(
            "serie",
            ""
        ).strip()

        club = request.form.get(
            "club",
            ""
        ).strip()

        estado = normalizar_estado(
            request.form.get(
                "estado",
                "Vigente"
            )
        )

        if not all([
            rut,
            nombre,
            fecha,
            serie,
            club
        ]):

            flash(
                "Completa todos los campos.",
                "error"
            )

            return render_template(
                "jugador_form.html",
                jugador=None,
                clubes=clubes,
                series=series
            )

        if not validar_rut(rut):

            flash(
                "El RUT ingresado no es válido.",
                "error"
            )

            return render_template(
                "jugador_form.html",
                jugador=None,
                clubes=clubes,
                series=series
            )

        try:

            fecha_obj = date.fromisoformat(
                fecha
            )

        except ValueError:

            flash(
                "Fecha no válida.",
                "error"
            )

            return render_template(
                "jugador_form.html",
                jugador=None,
                clubes=clubes,
                series=series
            )

        if Jugador.query.filter_by(
            rut=rut
        ).first():

            flash(
                "Ese RUT ya está registrado.",
                "error"
            )

            return render_template(
                "jugador_form.html",
                jugador=None,
                clubes=clubes,
                series=series
            )

        foto, error_foto = obtener_fotografia()

        if error_foto:

            flash(
                error_foto,
                "error"
            )

            return render_template(
                "jugador_form.html",
                jugador=None,
                clubes=clubes,
                series=series
            )

        jugador = Jugador(
            rut=rut,
            nombre_completo=nombre,
            fecha_nacimiento=fecha_obj,
            serie=serie,
            club=club,
            foto=foto,
            estado=estado
        )

        try:

            db.session.add(
                jugador
            )

            db.session.commit()

        except Exception as error:

            db.session.rollback()

            print(
                "Error registrando jugador:",
                error
            )

            flash(
                "No fue posible guardar el jugador.",
                "error"
            )

            return render_template(
                "jugador_form.html",
                jugador=None,
                clubes=clubes,
                series=series
            )

        flash(
            "Jugador registrado correctamente.",
            "success"
        )

        return redirect(
            url_for(
                "ficha_jugador",
                jugador_id=jugador.id
            )
        )

    return render_template(
        "jugador_form.html",
        jugador=None,
        clubes=clubes,
        series=series
    )


# ============================================================
# EDITAR JUGADOR
# ============================================================

@app.route(
    "/jugadores/<int:jugador_id>/editar",
    methods=["GET", "POST"]
)
def editar_jugador(jugador_id):

    jugador = db.get_or_404(
        Jugador,
        jugador_id
    )

    clubes, series = obtener_datos_formulario_jugador()

    if request.method == "POST":

        rut = normalizar_rut(
            request.form.get(
                "rut",
                ""
            )
        )

        nombre = request.form.get(
            "nombre_completo",
            ""
        ).strip()

        fecha = request.form.get(
            "fecha_nacimiento",
            ""
        ).strip()

        serie = request.form.get(
            "serie",
            ""
        ).strip()

        club = request.form.get(
            "club",
            ""
        ).strip()

        estado = normalizar_estado(
            request.form.get(
                "estado",
                jugador.estado or "Vigente"
            )
        )

        if not all([
            rut,
            nombre,
            fecha,
            serie,
            club
        ]):

            flash(
                "Completa todos los campos.",
                "error"
            )

            return render_template(
                "jugador_form.html",
                jugador=jugador,
                clubes=clubes,
                series=series
            )

        if not validar_rut(rut):

            flash(
                "El RUT ingresado no es válido.",
                "error"
            )

            return render_template(
                "jugador_form.html",
                jugador=jugador,
                clubes=clubes,
                series=series
            )

        otro_jugador = Jugador.query.filter(
            Jugador.rut == rut,
            Jugador.id != jugador.id
        ).first()

        if otro_jugador:

            flash(
                "Ese RUT ya pertenece a otro jugador.",
                "error"
            )

            return render_template(
                "jugador_form.html",
                jugador=jugador,
                clubes=clubes,
                series=series
            )

        try:

            fecha_obj = date.fromisoformat(
                fecha
            )

        except ValueError:

            flash(
                "Fecha no válida.",
                "error"
            )

            return render_template(
                "jugador_form.html",
                jugador=jugador,
                clubes=clubes,
                series=series
            )

        archivo_foto = request.files.get(
            "foto"
        )

        if (
            archivo_foto
            and archivo_foto.filename
        ):

            foto, error_foto = obtener_fotografia()

            if error_foto:

                flash(
                    error_foto,
                    "error"
                )

                return render_template(
                    "jugador_form.html",
                    jugador=jugador,
                    clubes=clubes,
                    series=series
                )

            jugador.foto = foto

        jugador.rut = rut

        jugador.nombre_completo = nombre

        jugador.fecha_nacimiento = fecha_obj

        jugador.serie = serie

        jugador.club = club

        jugador.estado = estado

        try:

            db.session.commit()

        except Exception as error:

            db.session.rollback()

            print(
                "Error actualizando jugador:",
                error
            )

            flash(
                "No fue posible actualizar el jugador.",
                "error"
            )

            return render_template(
                "jugador_form.html",
                jugador=jugador,
                clubes=clubes,
                series=series
            )

        flash(
            "Datos actualizados correctamente.",
            "success"
        )

        return redirect(
            url_for(
                "ficha_jugador",
                jugador_id=jugador.id
            )
        )

    return render_template(
        "jugador_form.html",
        jugador=jugador,
        clubes=clubes,
        series=series
    )


# ============================================================
# ELIMINAR JUGADOR
# ============================================================

@app.route(
    "/jugadores/<int:jugador_id>/eliminar",
    methods=["POST"]
)
def eliminar_jugador(jugador_id):

    jugador = db.get_or_404(
        Jugador,
        jugador_id
    )

    try:

        db.session.delete(
            jugador
        )

        db.session.commit()

    except Exception as error:

        db.session.rollback()

        print(
            "Error eliminando jugador:",
            error
        )

        flash(
            "No fue posible eliminar el jugador.",
            "error"
        )

        return redirect(
            url_for(
                "ficha_jugador",
                jugador_id=jugador_id
            )
        )

    flash(
        "Jugador eliminado.",
        "success"
    )

    return redirect(
        url_for("index")
    )


# ============================================================
# IMPORTAR EXCEL
# ============================================================

@app.route(
    "/jugadores/importar",
    methods=["GET", "POST"]
)
def importar_jugadores():

    if request.method == "GET":

        return render_template(
            "importar.html"
        )

    archivo = request.files.get(
        "archivo"
    )

    if not archivo or not archivo.filename:

        flash(
            "Selecciona un archivo Excel.",
            "error"
        )

        return redirect(
            url_for("importar_jugadores")
        )

    extension = (
        archivo.filename
        .rsplit(".", 1)[-1]
        .lower()
    )

    if extension != "xlsx":

        flash(
            "El archivo debe ser Excel (.xlsx).",
            "error"
        )

        return redirect(
            url_for("importar_jugadores")
        )

    try:

        workbook = load_workbook(
            archivo,
            data_only=True
        )

        hoja = workbook.active

        filas = list(
            hoja.iter_rows(
                values_only=True
            )
        )

        if not filas:

            flash(
                "El archivo está vacío.",
                "error"
            )

            return redirect(
                url_for("importar_jugadores")
            )

        encabezados = [
            str(x).strip().lower()
            if x is not None else ""
            for x in filas[0]
        ]

        columnas = {}

        equivalencias = {

            "rut": [
                "rut",
                "r.u.t.",
                "r.u.t",
                "documento"
            ],

            "nombre": [
                "nombre",
                "nombre completo",
                "nombre_completo"
            ],

            "fecha": [
                "fecha de nacimiento",
                "fecha nacimiento",
                "nacimiento",
                "fecha_nacimiento"
            ],

            "serie": [
                "serie",
                "categoría",
                "categoria"
            ],

            "club": [
                "club",
                "equipo"
            ]

        }

        for nombre_columna, posibles in (
            equivalencias.items()
        ):

            for posible in posibles:

                if posible in encabezados:

                    columnas[
                        nombre_columna
                    ] = encabezados.index(
                        posible
                    )

                    break

        faltantes = [
            campo
            for campo in equivalencias
            if campo not in columnas
        ]

        if faltantes:

            flash(
                "Faltan columnas obligatorias: "
                + ", ".join(faltantes),
                "error"
            )

            return redirect(
                url_for(
                    "importar_jugadores"
                )
            )

        registrados = 0

        duplicados = 0

        errores = 0

        detalle_errores = []

        ruts_archivo = set()

        for numero_fila, fila in enumerate(
            filas[1:],
            start=2
        ):

            try:

                rut_original = fila[
                    columnas["rut"]
                ]

                nombre = fila[
                    columnas["nombre"]
                ]

                fecha_valor = fila[
                    columnas["fecha"]
                ]

                serie = fila[
                    columnas["serie"]
                ]

                club = fila[
                    columnas["club"]
                ]

                rut = normalizar_rut(
                    rut_original
                )

                nombre = (
                    str(nombre).strip()
                    if nombre is not None
                    else ""
                )

                serie = (
                    str(serie).strip()
                    if serie is not None
                    else ""
                )

                club = (
                    str(club).strip()
                    if club is not None
                    else ""
                )

                if not all([
                    rut,
                    nombre,
                    fecha_valor,
                    serie,
                    club
                ]):

                    errores += 1

                    detalle_errores.append(
                        f"Fila {numero_fila}: "
                        "faltan datos obligatorios."
                    )

                    continue

                if not validar_rut(rut):

                    errores += 1

                    detalle_errores.append(
                        f"Fila {numero_fila}: "
                        f"RUT inválido ({rut})."
                    )

                    continue

                if rut in ruts_archivo:

                    duplicados += 1

                    continue

                ruts_archivo.add(rut)

                if Jugador.query.filter_by(
                    rut=rut
                ).first():

                    duplicados += 1

                    continue

                fecha_obj = convertir_fecha(
                    fecha_valor
                )

                if not fecha_obj:

                    errores += 1

                    detalle_errores.append(
                        f"Fila {numero_fila}: "
                        "fecha de nacimiento inválida."
                    )

                    continue

                jugador = Jugador(
                    rut=rut,
                    nombre_completo=nombre,
                    fecha_nacimiento=fecha_obj,
                    serie=serie,
                    club=club,
                    estado="Vigente"
                )

                db.session.add(
                    jugador
                )

                registrados += 1

            except Exception as error:

                errores += 1

                detalle_errores.append(
                    f"Fila {numero_fila}: "
                    f"error al procesar: {error}"
                )

        try:

            db.session.commit()

        except Exception:

            db.session.rollback()

            flash(
                "Ocurrió un error al guardar "
                "los jugadores.",
                "error"
            )

            return redirect(
                url_for(
                    "importar_jugadores"
                )
            )

        return render_template(
            "importar_resultado.html",
            registrados=registrados,
            duplicados=duplicados,
            errores=errores,
            detalle_errores=detalle_errores
        )

    except Exception as error:

        db.session.rollback()

        print(
            "Error importando Excel:",
            error
        )

        flash(
            "Ocurrió un error inesperado "
            "al importar el archivo.",
            "error"
        )

        return redirect(
            url_for(
                "importar_jugadores"
            )
        )


# ============================================================
# API DE JUGADORES
# ============================================================

@app.route(
    "/api/jugadores"
)
def api_jugadores():

    rut = request.args.get(
        "rut",
        ""
    ).strip().upper()

    if not rut:

        return jsonify([])

    jugadores = Jugador.query.filter(
        Jugador.rut.ilike(
            f"%{rut}%"
        )
    ).order_by(
        Jugador.nombre_completo
    ).all()

    return jsonify([

        {
            "id": jugador.id,

            "rut": jugador.rut,

            "nombre_completo":
                jugador.nombre_completo,

            "fecha_nacimiento":
                jugador.fecha_nacimiento.isoformat(),

            "serie":
                jugador.serie,

            "club":
                jugador.club,

            "estado":
                jugador.estado or "Vigente",

            "tiene_foto":
                bool(jugador.foto),

            "goles":
                obtener_goles(jugador.id),

            "amarillas":
                obtener_amarillas(jugador.id),

            "rojas":
                obtener_rojas(jugador.id),

            "suspensiones":
                obtener_suspensiones(jugador.id)
        }

        for jugador in jugadores

    ])


# ============================================================
# REGISTRAR GOL
# ============================================================

@app.route(
    "/jugadores/<int:jugador_id>/gol",
    methods=["POST"]
)
def registrar_gol(jugador_id):

    jugador = db.get_or_404(
        Jugador,
        jugador_id
    )

    try:

        cantidad = int(
            request.form.get(
                "cantidad",
                1
            )
        )

    except ValueError:

        cantidad = 1

    if cantidad < 1:
        cantidad = 1

    campeonato = request.form.get(
        "campeonato",
        ""
    ).strip()

    observaciones = request.form.get(
        "observaciones",
        ""
    ).strip()

    gol = Gol(
        jugador_id=jugador.id,
        fecha=date.today(),
        cantidad=cantidad,
        campeonato=campeonato,
        observaciones=observaciones
    )

    try:

        db.session.add(
            gol
        )

        db.session.commit()

    except Exception as error:

        db.session.rollback()

        print(
            "Error registrando gol:",
            error
        )

        flash(
            "No fue posible registrar el gol.",
            "error"
        )

        return redirect(
            url_for(
                "ficha_jugador",
                jugador_id=jugador.id
            )
        )

    flash(
        f"Se registraron {cantidad} gol(es).",
        "success"
    )

    return redirect(
        url_for(
            "ficha_jugador",
            jugador_id=jugador.id
        )
    )


# ============================================================
# REGISTRAR TARJETA AMARILLA
# ============================================================

@app.route(
    "/jugadores/<int:jugador_id>/amarilla",
    methods=["POST"]
)
def registrar_amarilla(jugador_id):

    jugador = db.get_or_404(
        Jugador,
        jugador_id
    )

    campeonato = request.form.get(
        "campeonato",
        ""
    ).strip()

    motivo = request.form.get(
        "motivo",
        ""
    ).strip()

    observaciones = request.form.get(
        "observaciones",
        ""
    ).strip()

    try:

        registro = RegistroDisciplinario(
            jugador_id=jugador.id,
            fecha=date.today(),
            tipo="Amarilla",
            cantidad=1,
            motivo=motivo,
            campeonato=campeonato,
            observaciones=observaciones
        )

        db.session.add(registro)

        db.session.flush()

        nuevas_suspensiones = crear_suspension_por_acumulacion(
            jugador,
            campeonato
        )

        db.session.commit()

    except Exception as error:

        db.session.rollback()

        print(
            "ERROR REGISTRANDO TARJETA AMARILLA:",
            repr(error)
        )

        flash(
            "No fue posible registrar la tarjeta amarilla.",
            "error"
        )

        return redirect(
            url_for(
                "ficha_jugador",
                jugador_id=jugador.id
            )
        )

    amarillas = obtener_amarillas(
        jugador.id
    )

    if nuevas_suspensiones:

        flash(
            f"Tarjeta amarilla registrada. "
            f"El jugador alcanzó {amarillas} amarillas "
            f"y se generó automáticamente una suspensión.",
            "warning"
        )

    else:

        flash(
            f"Tarjeta amarilla registrada correctamente. "
            f"Acumuladas: {amarillas}/4.",
            "success"
        )

    return redirect(
        url_for(
            "ficha_jugador",
            jugador_id=jugador.id
        )
    )


# ============================================================
# REGISTRAR TARJETA ROJA
# ============================================================

@app.route(
    "/jugadores/<int:jugador_id>/roja",
    methods=["POST"]
)
def registrar_roja(jugador_id):

    jugador = db.get_or_404(
        Jugador,
        jugador_id
    )

    campeonato = request.form.get(
        "campeonato",
        ""
    ).strip()

    motivo = request.form.get(
        "motivo",
        ""
    ).strip()

    observaciones = request.form.get(
        "observaciones",
        ""
    ).strip()

    try:

        registro = RegistroDisciplinario(
            jugador_id=jugador.id,
            fecha=date.today(),
            tipo="Roja",
            cantidad=1,
            motivo=motivo,
            campeonato=campeonato,
            observaciones=observaciones
        )

        db.session.add(registro)

        db.session.commit()

    except Exception as error:

        db.session.rollback()

        import traceback

        print("==============================================")
        print("ERROR REGISTRANDO TARJETA ROJA")
        print("==============================================")
        print(repr(error))
        traceback.print_exc()
        print("==============================================")

        flash(
            "No fue posible registrar la tarjeta roja.",
            "error"
        )

        return redirect(
            url_for(
                "ficha_jugador",
                jugador_id=jugador.id
            )
        )

    flash(
        "Tarjeta roja registrada correctamente.",
        "success"
    )

    return redirect(
        url_for(
            "ficha_jugador",
            jugador_id=jugador.id
        )
    )


# ============================================================
# REGISTRAR SUSPENSIÓN
# ============================================================

@app.route(
    "/jugadores/<int:jugador_id>/suspension",
    methods=["POST"]
)
def registrar_suspension(jugador_id):

    jugador = db.get_or_404(
        Jugador,
        jugador_id
    )

    try:

        cantidad = int(
            request.form.get(
                "cantidad",
                1
            )
        )

    except (ValueError, TypeError):

        cantidad = 1

    if cantidad < 1:
        cantidad = 1

    campeonato = request.form.get(
        "campeonato",
        ""
    ).strip()

    motivo = request.form.get(
        "motivo",
        ""
    ).strip()

    observaciones = request.form.get(
        "observaciones",
        ""
    ).strip()

    try:

        suspension = RegistroDisciplinario(
            jugador_id=jugador.id,
            fecha=date.today(),
            tipo="Suspension",
            cantidad=cantidad,
            motivo=motivo,
            campeonato=campeonato,
            observaciones=observaciones
        )

        db.session.add(suspension)

        jugador.estado = "Suspendido"

        db.session.commit()

    except Exception as error:

        db.session.rollback()

        print(
            "ERROR REGISTRANDO SUSPENSION:",
            repr(error)
        )

        flash(
            "No fue posible registrar la suspensión.",
            "error"
        )

        return redirect(
            url_for(
                "ficha_jugador",
                jugador_id=jugador.id
            )
        )

    flash(
        f"Suspensión registrada correctamente. "
        f"Cantidad: {cantidad}.",
        "success"
    )

    return redirect(
        url_for(
            "ficha_jugador",
            jugador_id=jugador.id
        )
    )

# ============================================================
# ELIMINAR REGISTRO DE GOL
# ============================================================

@app.route(
    "/jugadores/<int:jugador_id>/gol/<int:gol_id>/eliminar",
    methods=["POST"]
)
def eliminar_gol(jugador_id, gol_id):

    jugador = db.get_or_404(
        Jugador,
        jugador_id
    )

    gol = db.get_or_404(
        Gol,
        gol_id
    )

    if gol.jugador_id != jugador.id:

        flash(
            "El registro de gol no pertenece a este jugador.",
            "error"
        )

        return redirect(
            url_for(
                "ficha_jugador",
                jugador_id=jugador.id
            )
        )

    try:

        db.session.delete(gol)
        db.session.commit()

        flash(
            "Registro de gol eliminado correctamente.",
            "success"
        )

    except Exception as error:

        db.session.rollback()

        print("Error eliminando gol:", error)

        flash(
            "No fue posible eliminar el registro de gol.",
            "error"
        )

    return redirect(
        url_for(
            "ficha_jugador",
            jugador_id=jugador.id
        )
    )


# ============================================================
# ELIMINAR REGISTRO DISCIPLINARIO
# ============================================================

@app.route(
    "/jugadores/<int:jugador_id>/disciplina/<int:registro_id>/eliminar",
    methods=["POST"]
)
def eliminar_registro_disciplinario(jugador_id, registro_id):

    jugador = db.get_or_404(
        Jugador,
        jugador_id
    )

    registro = db.get_or_404(
        RegistroDisciplinario,
        registro_id
    )

    if registro.jugador_id != jugador.id:

        flash(
            "El registro disciplinario no pertenece a este jugador.",
            "error"
        )

        return redirect(
            url_for(
                "ficha_jugador",
                jugador_id=jugador.id
            )
        )

    try:

        db.session.delete(registro)
        db.session.commit()

        flash(
            "Registro disciplinario eliminado correctamente.",
            "success"
        )

    except Exception as error:

        db.session.rollback()

        print(
            "Error eliminando registro disciplinario:",
            error
        )

        flash(
            "No fue posible eliminar el registro disciplinario.",
            "error"
        )

    return redirect(
        url_for(
            "ficha_jugador",
            jugador_id=jugador.id
        )
    )


# ============================================================
# QR DEL JUGADOR
# ============================================================

@app.route(
    "/jugadores/<int:jugador_id>/qr"
)
def qr_jugador(jugador_id):

    jugador = db.get_or_404(
        Jugador,
        jugador_id
    )

    url_verificacion = url_for(
        "verificacion_jugador",
        jugador_id=jugador.id,
        _external=True
    )

    imagen_qr = qrcode.make(
        url_verificacion
    )

    memoria = BytesIO()

    imagen_qr.save(
        memoria,
        format="PNG"
    )

    memoria.seek(0)

    return Response(
        memoria.getvalue(),
        mimetype="image/png"
    )


# ============================================================
# VERIFICACIÓN PÚBLICA
# ============================================================

@app.route(
    "/jugadores/<int:jugador_id>/verificar"
)
def verificacion_jugador(jugador_id):

    jugador = db.get_or_404(
        Jugador,
        jugador_id
    )

    return render_template(
        "verificacion_jugador.html",
        jugador=jugador
    )


# ============================================================
# CREDENCIAL
# ============================================================

@app.route(
    "/jugadores/<int:jugador_id>/credencial"
)
def credencial_jugador(jugador_id):

    jugador = db.get_or_404(
        Jugador,
        jugador_id
    )

    return render_template(
        "jugador_credencial.html",
        jugador=jugador
    )


# ============================================================
# CREDENCIAL COMPLETA — FRENTE + REVERSO
# ============================================================

@app.route(
    "/jugadores/<int:jugador_id>/credencial-completa"
)
def credencial_completa(jugador_id):

    jugador = db.get_or_404(
        Jugador,
        jugador_id
    )

    return render_template(
        "credencial_completa.html",
        jugador=jugador
    )


# ============================================================
# REVERSO DE CREDENCIAL
# ============================================================

@app.route(
    "/jugadores/<int:jugador_id>/credencial/reverso"
)
def credencial_reverso(jugador_id):

    jugador = db.get_or_404(
        Jugador,
        jugador_id
    )

    return render_template(
        "credencial_reverso.html",
        jugador=jugador
    )


# ============================================================
# DASHBOARD
# ============================================================

@app.route("/dashboard")
def dashboard():

    # Filtro opcional por serie. Si viene vacío, muestra todas.
    serie_seleccionada = request.args.get("serie", "").strip()

    # Series disponibles para el selector.
    series_disponibles = (
        Serie.query
        .filter_by(activo=True)
        .order_by(Serie.nombre)
        .all()
    )

    # Consulta base de jugadores.
    jugadores_query = Jugador.query

    if serie_seleccionada:
        jugadores_query = jugadores_query.filter(
            Jugador.serie == serie_seleccionada
        )

    # ------------------------------------------------------------
    # INDICADORES GENERALES
    # ------------------------------------------------------------

    total_jugadores = jugadores_query.count()

    vigentes = jugadores_query.filter(
        Jugador.estado == "Vigente"
    ).count()

    pendientes = jugadores_query.filter(
        Jugador.estado == "Pendiente"
    ).count()

    suspendidos = jugadores_query.filter(
        Jugador.estado == "Suspendido"
    ).count()

    inhabilitados = jugadores_query.filter(
        Jugador.estado == "Inhabilitado"
    ).count()

    # ------------------------------------------------------------
    # CONSULTAS SEGURAS
    # ------------------------------------------------------------

    def safe_all(query, default=None):
        try:
            return query.all()
        except Exception as error:
            db.session.rollback()
            print("Advertencia Dashboard:", repr(error))
            return [] if default is None else default

    def safe_scalar(query, default=0):
        try:
            value = query.scalar()
            return value if value is not None else default
        except Exception as error:
            db.session.rollback()
            print("Advertencia Dashboard:", repr(error))
            return default

    # ------------------------------------------------------------
    # JUGADORES POR CLUB
    # ------------------------------------------------------------

    query_club = db.session.query(
        Jugador.club,
        db.func.count(Jugador.id)
    ).filter(
        Jugador.club.isnot(None),
        Jugador.club != ""
    )

    if serie_seleccionada:
        query_club = query_club.filter(
            Jugador.serie == serie_seleccionada
        )

    jugadores_por_club = safe_all(
        query_club
        .group_by(Jugador.club)
        .order_by(db.func.count(Jugador.id).desc())
    )

    # ------------------------------------------------------------
    # JUGADORES POR SERIE
    # ------------------------------------------------------------

    query_serie = db.session.query(
        Jugador.serie,
        db.func.count(Jugador.id)
    ).filter(
        Jugador.serie.isnot(None),
        Jugador.serie != ""
    )

    if serie_seleccionada:
        query_serie = query_serie.filter(
            Jugador.serie == serie_seleccionada
        )

    jugadores_por_serie = safe_all(
        query_serie
        .group_by(Jugador.serie)
        .order_by(db.func.count(Jugador.id).desc())
    )

    # ------------------------------------------------------------
    # ÚLTIMOS JUGADORES
    # ------------------------------------------------------------

    ultimos_query = jugadores_query.order_by(
        Jugador.id.desc()
    )

    ultimos_jugadores = ultimos_query.limit(5).all()

    # ------------------------------------------------------------
    # ESTADÍSTICAS DEPORTIVAS FILTRADAS POR SERIE
    # ------------------------------------------------------------

    query_goles = db.session.query(
        db.func.coalesce(db.func.sum(Gol.cantidad), 0)
    ).join(
        Jugador,
        Gol.jugador_id == Jugador.id
    )

    query_amarillas = db.session.query(
        db.func.coalesce(db.func.sum(RegistroDisciplinario.cantidad), 0)
    ).join(
        Jugador,
        RegistroDisciplinario.jugador_id == Jugador.id
    ).filter(
        RegistroDisciplinario.tipo == "Amarilla"
    )

    query_rojas = db.session.query(
        db.func.coalesce(db.func.sum(RegistroDisciplinario.cantidad), 0)
    ).join(
        Jugador,
        RegistroDisciplinario.jugador_id == Jugador.id
    ).filter(
        RegistroDisciplinario.tipo == "Roja"
    )

    query_suspensiones = db.session.query(
        db.func.coalesce(db.func.sum(RegistroDisciplinario.cantidad), 0)
    ).join(
        Jugador,
        RegistroDisciplinario.jugador_id == Jugador.id
    ).filter(
        RegistroDisciplinario.tipo == "Suspension"
    )

    if serie_seleccionada:
        query_goles = query_goles.filter(Jugador.serie == serie_seleccionada)
        query_amarillas = query_amarillas.filter(Jugador.serie == serie_seleccionada)
        query_rojas = query_rojas.filter(Jugador.serie == serie_seleccionada)
        query_suspensiones = query_suspensiones.filter(Jugador.serie == serie_seleccionada)

    total_goles = safe_scalar(query_goles)
    total_amarillas = safe_scalar(query_amarillas)
    total_rojas = safe_scalar(query_rojas)
    total_suspensiones = safe_scalar(query_suspensiones)

    # ------------------------------------------------------------
    # INDICADORES DE PARTICIPACIÓN DEPORTIVA
    # ------------------------------------------------------------

    query_jugadores_goles = db.session.query(
        db.func.count(db.func.distinct(Gol.jugador_id))
    ).join(Jugador, Gol.jugador_id == Jugador.id)

    query_jugadores_amarillas = db.session.query(
        db.func.count(db.func.distinct(RegistroDisciplinario.jugador_id))
    ).join(Jugador, RegistroDisciplinario.jugador_id == Jugador.id).filter(
        RegistroDisciplinario.tipo == "Amarilla"
    )

    query_jugadores_rojas = db.session.query(
        db.func.count(db.func.distinct(RegistroDisciplinario.jugador_id))
    ).join(Jugador, RegistroDisciplinario.jugador_id == Jugador.id).filter(
        RegistroDisciplinario.tipo == "Roja"
    )

    query_jugadores_suspendidos = db.session.query(
        db.func.count(db.func.distinct(RegistroDisciplinario.jugador_id))
    ).join(Jugador, RegistroDisciplinario.jugador_id == Jugador.id).filter(
        RegistroDisciplinario.tipo == "Suspension"
    )

    if serie_seleccionada:
        query_jugadores_goles = query_jugadores_goles.filter(Jugador.serie == serie_seleccionada)
        query_jugadores_amarillas = query_jugadores_amarillas.filter(Jugador.serie == serie_seleccionada)
        query_jugadores_rojas = query_jugadores_rojas.filter(Jugador.serie == serie_seleccionada)
        query_jugadores_suspendidos = query_jugadores_suspendidos.filter(Jugador.serie == serie_seleccionada)

    jugadores_con_goles = safe_scalar(query_jugadores_goles)
    jugadores_con_amarillas = safe_scalar(query_jugadores_amarillas)
    jugadores_con_rojas = safe_scalar(query_jugadores_rojas)
    jugadores_suspendidos_registro = safe_scalar(query_jugadores_suspendidos)

    # ------------------------------------------------------------
    # GOLEADORES
    # ------------------------------------------------------------

    query_goleadores = db.session.query(
        Jugador.id,
        Jugador.nombre_completo,
        Jugador.club,
        Jugador.serie,
        db.func.sum(Gol.cantidad).label("total")
    ).join(
        Gol,
        Gol.jugador_id == Jugador.id
    )

    if serie_seleccionada:
        query_goleadores = query_goleadores.filter(
            Jugador.serie == serie_seleccionada
        )

    goleadores = safe_all(
        query_goleadores
        .group_by(
            Jugador.id,
            Jugador.nombre_completo,
            Jugador.club,
            Jugador.serie
        )
        .order_by(
            db.func.sum(Gol.cantidad).desc(),
            Jugador.nombre_completo.asc()
        )
        .limit(10)
    )

    # ------------------------------------------------------------
    # RANKING AMARILLAS
    # ------------------------------------------------------------

    query_ranking_amarillas = db.session.query(
        Jugador.id,
        Jugador.nombre_completo,
        Jugador.club,
        Jugador.serie,
        db.func.sum(RegistroDisciplinario.cantidad).label("total")
    ).join(
        RegistroDisciplinario,
        RegistroDisciplinario.jugador_id == Jugador.id
    ).filter(
        RegistroDisciplinario.tipo == "Amarilla"
    )

    if serie_seleccionada:
        query_ranking_amarillas = query_ranking_amarillas.filter(
            Jugador.serie == serie_seleccionada
        )

    ranking_amarillas = safe_all(
        query_ranking_amarillas
        .group_by(
            Jugador.id,
            Jugador.nombre_completo,
            Jugador.club,
            Jugador.serie
        )
        .order_by(
            db.func.sum(RegistroDisciplinario.cantidad).desc(),
            Jugador.nombre_completo.asc()
        )
        .limit(10)
    )

    # ------------------------------------------------------------
    # RANKING ROJAS
    # ------------------------------------------------------------

    query_ranking_rojas = db.session.query(
        Jugador.id,
        Jugador.nombre_completo,
        Jugador.club,
        Jugador.serie,
        db.func.sum(RegistroDisciplinario.cantidad).label("total")
    ).join(
        RegistroDisciplinario,
        RegistroDisciplinario.jugador_id == Jugador.id
    ).filter(
        RegistroDisciplinario.tipo == "Roja"
    )

    if serie_seleccionada:
        query_ranking_rojas = query_ranking_rojas.filter(
            Jugador.serie == serie_seleccionada
        )

    ranking_rojas = safe_all(
        query_ranking_rojas
        .group_by(
            Jugador.id,
            Jugador.nombre_completo,
            Jugador.club,
            Jugador.serie
        )
        .order_by(
            db.func.sum(RegistroDisciplinario.cantidad).desc(),
            Jugador.nombre_completo.asc()
        )
        .limit(10)
    )

    # ------------------------------------------------------------
    # RANKING SUSPENSIONES
    # ------------------------------------------------------------

    query_ranking_suspensiones = db.session.query(
        Jugador.id,
        Jugador.nombre_completo,
        Jugador.club,
        Jugador.serie,
        db.func.sum(RegistroDisciplinario.cantidad).label("total")
    ).join(
        RegistroDisciplinario,
        RegistroDisciplinario.jugador_id == Jugador.id
    ).filter(
        RegistroDisciplinario.tipo == "Suspension"
    )

    if serie_seleccionada:
        query_ranking_suspensiones = query_ranking_suspensiones.filter(
            Jugador.serie == serie_seleccionada
        )

    ranking_suspensiones = safe_all(
        query_ranking_suspensiones
        .group_by(
            Jugador.id,
            Jugador.nombre_completo,
            Jugador.club,
            Jugador.serie
        )
        .order_by(
            db.func.sum(RegistroDisciplinario.cantidad).desc(),
            Jugador.nombre_completo.asc()
        )
        .limit(10)
    )

    # ------------------------------------------------------------
    # COMPARATIVA GENERAL POR SERIE
    # ------------------------------------------------------------

    # Esta tabla se calcula para todas las series activas,
    # independiente del filtro actualmente seleccionado.
    estadisticas_por_serie = []

    for serie_item in series_disponibles:

        nombre_serie = serie_item.nombre

        jugadores_serie = Jugador.query.filter(
            Jugador.serie == nombre_serie
        )

        goles_serie = safe_scalar(
            db.session.query(
                db.func.coalesce(db.func.sum(Gol.cantidad), 0)
            ).join(
                Jugador,
                Gol.jugador_id == Jugador.id
            ).filter(
                Jugador.serie == nombre_serie
            )
        )

        amarillas_serie = safe_scalar(
            db.session.query(
                db.func.coalesce(
                    db.func.sum(RegistroDisciplinario.cantidad),
                    0
                )
            ).join(
                Jugador,
                RegistroDisciplinario.jugador_id == Jugador.id
            ).filter(
                Jugador.serie == nombre_serie,
                RegistroDisciplinario.tipo == "Amarilla"
            )
        )

        rojas_serie = safe_scalar(
            db.session.query(
                db.func.coalesce(
                    db.func.sum(RegistroDisciplinario.cantidad),
                    0
                )
            ).join(
                Jugador,
                RegistroDisciplinario.jugador_id == Jugador.id
            ).filter(
                Jugador.serie == nombre_serie,
                RegistroDisciplinario.tipo == "Roja"
            )
        )

        suspensiones_serie = safe_scalar(
            db.session.query(
                db.func.coalesce(
                    db.func.sum(RegistroDisciplinario.cantidad),
                    0
                )
            ).join(
                Jugador,
                RegistroDisciplinario.jugador_id == Jugador.id
            ).filter(
                Jugador.serie == nombre_serie,
                RegistroDisciplinario.tipo == "Suspension"
            )
        )

        estadisticas_por_serie.append({
            "serie": nombre_serie,
            "jugadores": jugadores_serie.count(),
            "vigentes": jugadores_serie.filter(
                Jugador.estado == "Vigente"
            ).count(),
            "suspendidos": jugadores_serie.filter(
                Jugador.estado == "Suspendido"
            ).count(),
            "goles": goles_serie,
            "amarillas": amarillas_serie,
            "rojas": rojas_serie,
            "suspensiones": suspensiones_serie
        })

    estadisticas_por_serie.sort(
        key=lambda item: item["jugadores"],
        reverse=True
    )

    # ------------------------------------------------------------
    # SUSPENDIDOS ACTUALMENTE
    # ------------------------------------------------------------

    suspendidos_query = jugadores_query.filter(
        Jugador.estado == "Suspendido"
    ).order_by(Jugador.nombre_completo.asc())

    jugadores_suspendidos = suspendidos_query.limit(10).all()

    return render_template(
        "dashboard.html",
        total_jugadores=total_jugadores,
        vigentes=vigentes,
        pendientes=pendientes,
        suspendidos=suspendidos,
        inhabilitados=inhabilitados,
        jugadores_por_club=jugadores_por_club,
        jugadores_por_serie=jugadores_por_serie,
        ultimos_jugadores=ultimos_jugadores,
        total_goles=total_goles,
        total_amarillas=total_amarillas,
        total_rojas=total_rojas,
        total_suspensiones=total_suspensiones,
        jugadores_con_goles=jugadores_con_goles,
        jugadores_con_amarillas=jugadores_con_amarillas,
        jugadores_con_rojas=jugadores_con_rojas,
        jugadores_suspendidos_registro=jugadores_suspendidos_registro,
        goleadores=goleadores,
        ranking_amarillas=ranking_amarillas,
        ranking_rojas=ranking_rojas,
        ranking_suspensiones=ranking_suspensiones,
        jugadores_suspendidos=jugadores_suspendidos,
        estadisticas_por_serie=estadisticas_por_serie,
        series_disponibles=series_disponibles,
        serie_seleccionada=serie_seleccionada
    )


# ============================================================
# ADMINISTRACIÓN DE CLUBES Y SERIES
# ============================================================

@app.route(
    "/configuracion"
)
def configuracion():

    clubes = Club.query.order_by(
        Club.nombre
    ).all()

    series = Serie.query.order_by(
        Serie.nombre
    ).all()

    return render_template(
        "configuracion.html",
        clubes=clubes,
        series=series
    )


# ============================================================
# CREAR CLUB
# ============================================================

@app.route(
    "/clubes/nuevo",
    methods=["POST"]
)
def nuevo_club():

    nombre = request.form.get(
        "nombre",
        ""
    ).strip()

    if not nombre:

        flash(
            "Debes ingresar el nombre del club.",
            "error"
        )

        return redirect(
            url_for("configuracion")
        )

    club_existente = Club.query.filter(
        db.func.lower(Club.nombre) ==
        nombre.lower()
    ).first()

    if club_existente:

        flash(
            "Ese club ya está registrado.",
            "error"
        )

        return redirect(
            url_for("configuracion")
        )

    club = Club(
        nombre=nombre,
        activo=True
    )

    db.session.add(
        club
    )

    db.session.commit()

    flash(
        f"Club '{nombre}' agregado correctamente.",
        "success"
    )

    return redirect(
        url_for("configuracion")
    )


# ============================================================
# CREAR SERIE
# ============================================================

@app.route(
    "/series/nueva",
    methods=["POST"]
)
def nueva_serie():

    nombre = request.form.get(
        "nombre",
        ""
    ).strip()

    if not nombre:

        flash(
            "Debes ingresar el nombre de la serie.",
            "error"
        )

        return redirect(
            url_for("configuracion")
        )

    serie_existente = Serie.query.filter(
        db.func.lower(Serie.nombre) ==
        nombre.lower()
    ).first()

    if serie_existente:

        flash(
            "Esa serie ya está registrada.",
            "error"
        )

        return redirect(
            url_for("configuracion")
        )

    serie = Serie(
        nombre=nombre,
        activo=True
    )

    db.session.add(
        serie
    )

    db.session.commit()

    flash(
        f"Serie '{nombre}' agregada correctamente.",
        "success"
    )

    return redirect(
        url_for("configuracion")
    )


# ============================================================
# ACTIVAR / DESACTIVAR CLUB
# ============================================================

@app.route(
    "/clubes/<int:club_id>/estado",
    methods=["POST"]
)
def cambiar_estado_club(club_id):

    club = db.get_or_404(
        Club,
        club_id
    )

    club.activo = not club.activo

    db.session.commit()

    estado = (
        "activado"
        if club.activo
        else "desactivado"
    )

    flash(
        f"Club {estado} correctamente.",
        "success"
    )

    return redirect(
        url_for("configuracion")
    )


# ============================================================
# ACTIVAR / DESACTIVAR SERIE
# ============================================================

@app.route(
    "/series/<int:serie_id>/estado",
    methods=["POST"]
)
def cambiar_estado_serie(serie_id):

    serie = db.get_or_404(
        Serie,
        serie_id
    )

    serie.activo = not serie.activo

    db.session.commit()

    estado = (
        "activada"
        if serie.activo
        else "desactivada"
    )

    flash(
        f"Serie {estado} correctamente.",
        "success"
    )

    return redirect(
        url_for("configuracion")
    )


# ============================================================
# HEALTH CHECK
# ============================================================

@app.route(
    "/health"
)
def health():

    return {
        "status": "ok"
    }


# ============================================================
# EJECUCIÓN LOCAL
# ============================================================

# ============================================================
# MÓDULO DE CAMPEONATOS — PASOS 1, 2 Y 3
# ============================================================

@app.route("/campeonatos")
def campeonatos():

    campeonatos_registrados = (
        Campeonato.query
        .order_by(
            Campeonato.temporada.desc(),
            Campeonato.id.desc()
        )
        .all()
    )

    return render_template(
        "campeonatos.html",
        campeonatos=campeonatos_registrados
    )


@app.route("/campeonatos/nuevo", methods=["GET", "POST"])
def nuevo_campeonato():

    series = (
        Serie.query
        .filter_by(activo=True)
        .order_by(Serie.nombre)
        .all()
    )

    if request.method == "POST":

        nombre = request.form.get("nombre", "").strip()
        temporada = request.form.get("temporada", "").strip()
        serie = request.form.get("serie", "").strip()
        estado = request.form.get("estado", "Activo").strip() or "Activo"
        descripcion = request.form.get("descripcion", "").strip()

        fecha_inicio = convertir_fecha(
            request.form.get("fecha_inicio", "").strip()
        )

        fecha_termino = convertir_fecha(
            request.form.get("fecha_termino", "").strip()
        )

        if not nombre or not temporada or not serie:
            flash(
                "Debes completar nombre, temporada y serie.",
                "error"
            )
            return render_template(
                "campeonato_form.html",
                series=series
            )

        if estado not in {"Activo", "Finalizado"}:
            estado = "Activo"

        if fecha_inicio and fecha_termino and fecha_termino < fecha_inicio:
            flash(
                "La fecha de término no puede ser anterior a la fecha de inicio.",
                "error"
            )
            return render_template(
                "campeonato_form.html",
                series=series
            )

        campeonato = Campeonato(
            nombre=nombre,
            temporada=temporada,
            serie=serie,
            fecha_inicio=fecha_inicio,
            fecha_termino=fecha_termino,
            estado=estado,
            descripcion=descripcion
        )

        try:
            db.session.add(campeonato)
            db.session.commit()

        except Exception as error:
            db.session.rollback()
            print(
                "ERROR CREANDO CAMPEONATO:",
                repr(error)
            )
            flash(
                "No fue posible crear el campeonato. Revise el registro del servidor.",
                "error"
            )
            return render_template(
                "campeonato_form.html",
                series=series
            )

        flash(
            f"Campeonato '{nombre}' creado correctamente.",
            "success"
        )

        return redirect(
            url_for(
                "detalle_campeonato",
                campeonato_id=campeonato.id
            )
        )

    return render_template(
        "campeonato_form.html",
        series=series
    )


@app.route("/campeonatos/<int:campeonato_id>")
def detalle_campeonato(campeonato_id):

    campeonato = db.get_or_404(
        Campeonato,
        campeonato_id
    )

    return render_template(
        "campeonato_detalle.html",
        campeonato=campeonato
    )



if __name__ == "__main__":

    app.run(
        host="0.0.0.0",
        port=int(
            os.environ.get(
                "PORT",
                5000
            )
        )
    )
